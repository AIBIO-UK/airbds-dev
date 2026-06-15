#!/usr/bin/env python3
"""Build the canonical metric YAML and CSV from the scoring spreadsheet.

The YAML and CSV are a coupled pair (see metric/README.md) and the
metric-alignment-check workflow requires them to change together. This script
regenerates BOTH from the same source so they can never drift apart.

Usage:
    # Regenerate metric/airbds_metric_v0.3.{yaml,csv} from the source workbook
    python3 src/scripts/build_metric_yaml_and_csv_from_spreadsheet_v0.3.py

    # Verify both committed files are in sync with the spreadsheet (CI-friendly)
    python3 src/scripts/build_metric_yaml_and_csv_from_spreadsheet_v0.3.py --check

    # Custom paths
    python3 src/scripts/build_metric_yaml_and_csv_from_spreadsheet_v0.3.py \
        --workbook "metric/source/AIRBDS Core Metric scoring v0.3 - _initials_-_#_ TEMPLATE.xlsx" \
        --output metric/airbds_metric_v0.3.yaml \
        --output-csv metric/airbds_metric_v0.3.csv

What comes from where
    The spreadsheet is the source of truth for everything that varies per
    question or per grade:
      • Scoring sheet  → the 28 questions (scope, theme, grade, mapped_from,
                          question text, guidance).
      • Lookups sheet  → grade_points and the grading thresholds
                          (min_proportion_yes + min_score, with descriptions).
    Document-level metadata that does NOT live in the spreadsheet (licence,
    repository, contact, the prose description, scope descriptions, and the
    banner comments) is held in the CONFIG block below — edit it there. CSV is a
    flat format with no comments, so the YAML's generated-file banner has no CSV
    counterpart; the pairing rule is satisfied because both files regenerate
    together from this script.

    The QUESTION SUMMARY footer is computed from the question table, so it can
    never drift out of sync with the questions above it.

Requires: openpyxl, PyYAML.

Exit codes:
    0 — wrote the files (default), or --check found both already in sync
    1 — --check found a committed file differs from the spreadsheet
    2 — the spreadsheet failed a sanity check (e.g. an unexpected scope)
"""

import argparse
import csv
import io
import re
import sys
from pathlib import Path

import openpyxl
import yaml

# Repo root is three levels up: <root>/src/scripts/<this file>.
REPO_ROOT = Path(__file__).resolve().parent.parent.parent

# This script's own path (for the drift-message hint), derived dynamically so it
# stays correct even if the file is renamed.
SCRIPT_PATH = f"src/scripts/{Path(__file__).name}"

# ── Editorial config (NOT in the spreadsheet — edit here) ────────────────────

VERSION = "0.3"

DEFAULT_WORKBOOK = REPO_ROOT / "metric" / "source" / "AIRBDS Core Metric scoring v0.3 - _initials_-_#_ TEMPLATE.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "metric" / "airbds_metric_v0.3.yaml"
DEFAULT_OUTPUT_CSV = REPO_ROOT / "metric" / "airbds_metric_v0.3.csv"

# Column order of the companion CSV. "weight"/"weight_points" mirror the YAML's
# grade/grade_points; mapped_from is the verbatim "A / B" spreadsheet form.
CSV_COLUMNS = [
    "question_id", "scope", "theme", "weight", "weight_points",
    "mapped_from", "question", "guidance", "not_applicable_default",
]

METADATA = {
    "schema_version": VERSION,
    "metric_name": "AIRBDS AI-Readiness Dataset Scoring Metric",
    "short_name": "AIRBDS Metric",
    "version": VERSION,
    "release_date": "2025",
    "license": "CC-BY-4.0",
    "repository": "https://github.com/AIBIO-UK/airbds-metric",
    "contact": "info@aibio.ac.uk",
}

DESCRIPTION_LINES = [
    "A structured checklist for evaluating the AI-readiness of bioscience datasets.",
    "Questions are grouped into four scopes (Infrastructure, Metadata, Content, Ethics)",
    "and graded as Optional, Important, or Critical to produce a composite score",
    "and grade (Caution / Bronze / Silver / Gold).",
]

# Scope order (sort_order is the 1-based position) and descriptions. The set of
# scope ids here is validated against the scopes actually used in the workbook.
SCOPES = [
    ("Infrastructure", "Questions about how the dataset is hosted, accessed, and identified."),
    ("Metadata", "Questions about the dataset's descriptive and contextual metadata."),
    ("Content", "Questions about the quality, format, and consistency of the data itself."),
    ("Ethics", "Questions about ethical, privacy, and security considerations. Applicable primarily to datasets containing human or animal subject data."),
]

# Scopes whose questions carry not_applicable_default: "Yes".
NA_DEFAULT_SCOPES = {"Ethics"}

QID_RE = re.compile(r"^ACM-\d+$")

# Banner / section comments, reproduced verbatim.
HEADER_COMMENT = f"""\
# =============================================================================
# {METADATA["metric_name"]} — v{VERSION}
# AI-Ready Bioscience Datasets Working Group (AIRBDS), AIBIO-UK
# https://aibio.ac.uk/about/working-groups/airbds/
# License: {METADATA["license"]}
# =============================================================================
#
# *** GENERATED FILE — DO NOT EDIT BY HAND. ***
#
# This file is generated from the AIRBDS scoring spreadsheet by the metric
# build script (see metric/README.md). Any manual edit will be lost the next
# time the file is regenerated. To change this file's contents — or the way it
# is generated — edit that script (and/or the source spreadsheet) and re-run it.
# Never edit this file directly.
# =============================================================================
#
# Question definitions for a given metric version. For a version the scope,
# theme and grade of each question are fixed, so they are defined here as the
# source of truth rather than trusted from uploaded assessments.
#
# This is a language-neutral data file: the question definitions can be read by
# any tool (the review processor, the auto-airbds web frontend, etc.). The
# `guidance` field explains how each question should be answered; it is guidance
# for humans/LLMs."""

GRADE_POINTS_COMMENT = """\
# Full points awarded for a question when its answer is "Yes", keyed by the
# question's grade. A "No" answer always scores 0. A question's score is
# therefore fixed by its grade and answer, not taken from the assessment."""

GRADING_COMMENT = """\
# Grading thresholds, highest grade first. A dataset earns the highest grade for
# which every requirement holds: the proportion of "Yes" answers within each
# grade category must be at least min_proportion_yes (compared with >=), and the
# total score must be at least min_score. The final entry is the floor (all-zero
# requirements). Editing these values re-grades without any code change."""

SCOPES_COMMENT = "# Top-level groupings. Scope is broader than theme."

QUESTIONS_COMMENT = """\
# Every question is a boolean "Yes"/"No". Ethics questions carry a
# not_applicable_default of "Yes" — they are marked "Yes" for datasets with no
# human or animal subject data."""


# ── Helpers ──────────────────────────────────────────────────────────────────

def dq(text: str) -> str:
    """Render a value as a YAML double-quoted scalar."""
    escaped = str(text).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def norm(text) -> str:
    """Collapse internal whitespace runs and strip (fixes stray spreadsheet spaces)."""
    return " ".join(str(text or "").split())


def fmt_prop(value) -> str:
    """Format a proportion the way the YAML does: 1.0, 0.5, 0.875, 0.0."""
    return str(float(value))


# ── Spreadsheet readers ──────────────────────────────────────────────────────

def read_questions(ws) -> list[dict]:
    """Read the question table from the Scoring sheet, in sheet order."""
    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    required = ["Q ID", "Mapped from", "Theme", "Scope", "Question", "Notes", "Grade"]
    missing = [h for h in required if h not in header]
    if missing:
        sys.exit(f"ERROR: Scoring sheet is missing column(s): {missing}")

    questions = []
    for r in range(2, ws.max_row + 1):
        qid = norm(ws.cell(r, header["Q ID"]).value)
        if not QID_RE.match(qid):
            continue
        mapped_raw = ws.cell(r, header["Mapped from"]).value
        mapped_from = [s.strip() for s in str(mapped_raw).split("/") if s.strip()] if mapped_raw else []
        questions.append({
            "id": qid,
            "scope": norm(ws.cell(r, header["Scope"]).value),
            "theme": norm(ws.cell(r, header["Theme"]).value),
            "grade": norm(ws.cell(r, header["Grade"]).value),
            "mapped_from": mapped_from,
            "question": norm(ws.cell(r, header["Question"]).value),
            "guidance": norm(ws.cell(r, header["Notes"]).value),
        })
    return questions


def _find_row(ws, predicate):
    for r in range(1, ws.max_row + 1):
        if predicate(r):
            return r
    return None


def read_grade_points(ws) -> dict:
    """Read {grade: points} from the Lookups 'Grade / Points' table."""
    hdr = _find_row(ws, lambda r: norm(ws.cell(r, 1).value) == "Grade"
                    and norm(ws.cell(r, 2).value) == "Points")
    if hdr is None:
        sys.exit("ERROR: could not locate the 'Grade / Points' table in Lookups")
    points = {}
    for r in range(hdr + 1, ws.max_row + 1):
        name = norm(ws.cell(r, 1).value)
        val = ws.cell(r, 2).value
        if name in {"Optional", "Important", "Critical"} and isinstance(val, (int, float)):
            points[name] = int(val)
        elif name or val is not None:
            break
    return points


def read_grading(ws) -> list[dict]:
    """Read the grading thresholds from the Lookups 'Required proportions' table."""
    hdr = _find_row(ws, lambda r: norm(ws.cell(r, 1).value) == "Optional"
                    and norm(ws.cell(r, 4).value) == "Threshold"
                    and norm(ws.cell(r, 5).value) == "Grade")
    if hdr is None:
        sys.exit("ERROR: could not locate the 'Required proportions' table in Lookups")
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        grade = norm(ws.cell(r, 5).value)
        if not grade:
            break
        rows.append({
            "name": grade,
            "description": norm(ws.cell(r, 6).value),
            "min_proportion_yes": {
                "Critical": float(ws.cell(r, 3).value or 0),
                "Important": float(ws.cell(r, 2).value or 0),
                "Optional": float(ws.cell(r, 1).value or 0),
            },
            "min_score": int(ws.cell(r, 4).value or 0),
        })
    # Highest grade first.
    rows.sort(key=lambda g: g["min_score"], reverse=True)
    return rows


# ── YAML rendering ─────────────────────────────────────────────────────────────

def render_footer(questions: list[dict], grade_points: dict) -> str:
    """Compute the QUESTION SUMMARY banner from the question table."""
    def num(qid):
        return int(qid.split("-")[1])

    lines = [
        "# =============================================================================",
        "# QUESTION SUMMARY",
        f"# Total questions: {len(questions)}",
    ]
    for scope_id, _ in SCOPES:
        ids = [q["id"] for q in questions if q["scope"] == scope_id]
        if not ids:
            continue
        ids.sort(key=num)
        rng = f"({ids[0]:<7}– {ids[-1]})"
        lines.append(f"# {(scope_id + ':').ljust(17)}{len(ids):>2}  {rng}")

    lines.append("#")
    lines.append(f"# Grade breakdown (all {len(questions)} questions):")
    total = 0
    for grade in ("Critical", "Important", "Optional"):
        count = sum(1 for q in questions if q["grade"] == grade)
        pts = grade_points[grade]
        subtotal = count * pts
        total += subtotal
        lines.append(
            f"#   {(grade + ':').ljust(10)}{count:>3}   × {str(pts):<2} pts = {str(subtotal):<3} pts maximum"
        )
    lines.append(f"#   TOTAL maximum (incl. Ethics): {total} pts")
    lines.append("# =============================================================================")
    return "\n".join(lines)


def render_yaml(questions, grade_points, grading) -> str:
    out = [HEADER_COMMENT, ""]

    for key, val in METADATA.items():
        out.append(f"{key}: {dq(val)}")
    out.append("")

    out.append("description: >")
    out.extend(f"  {line}" for line in DESCRIPTION_LINES)
    out.append("")

    out.append(GRADE_POINTS_COMMENT)
    out.append("grade_points:")
    for grade in sorted(grade_points, key=grade_points.get, reverse=True):
        out.append(f"  {grade}: {grade_points[grade]}")
    out.append("")

    out.append(GRADING_COMMENT)
    out.append("grading:")
    for g in grading:
        p = g["min_proportion_yes"]
        out.append(f"  - name: {g['name']}")
        out.append(f"    description: {dq(g['description'])}")
        out.append(
            "    min_proportion_yes: "
            f"{{ Critical: {fmt_prop(p['Critical'])}, "
            f"Important: {fmt_prop(p['Important'])}, "
            f"Optional: {fmt_prop(p['Optional'])} }}"
        )
        out.append(f"    min_score: {g['min_score']}")
    out.append("")

    out.append(SCOPES_COMMENT)
    out.append("scopes:")
    for i, (scope_id, desc) in enumerate(SCOPES, start=1):
        out.append(f"  - id: {scope_id}")
        out.append(f"    sort_order: {i}")
        out.append(f"    description: {dq(desc)}")
    out.append("")

    out.append(QUESTIONS_COMMENT)
    out.append("questions:")
    for q in questions:
        mapped = "[" + ", ".join(dq(m) for m in q["mapped_from"]) + "]"
        out.append(f"  {q['id']}:")
        out.append(f"    scope: {q['scope']}")
        out.append(f"    theme: {q['theme']}")
        out.append(f"    grade: {q['grade']}")
        out.append(f"    mapped_from: {mapped}")
        out.append(f"    question: {dq(q['question'])}")
        out.append(f"    guidance: {dq(q['guidance'])}")
        out.append('    answer: { type: boolean, options: ["Yes", "No"] }')
        if q["scope"] in NA_DEFAULT_SCOPES:
            out.append('    not_applicable_default: "Yes"')
    out.append("")

    out.append(render_footer(questions, grade_points))
    return "\n".join(out) + "\n"


def render_csv(questions, grade_points) -> str:
    """Render the companion CSV (CRLF, minimal quoting) from the same data."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(CSV_COLUMNS)
    for q in questions:
        writer.writerow([
            q["id"],
            q["scope"],
            q["theme"],
            q["grade"],
            grade_points[q["grade"]],
            " / ".join(q["mapped_from"]),
            q["question"],
            q["guidance"],
            "Yes" if q["scope"] in NA_DEFAULT_SCOPES else "",
        ])
    return buf.getvalue()


# ── Validation ─────────────────────────────────────────────────────────────────

def validate(questions, grade_points, grading) -> None:
    if not questions:
        sys.exit("ERROR: no questions (ACM-N rows) were found in the Scoring sheet")
    configured_scopes = {s for s, _ in SCOPES}
    used_scopes = {q["scope"] for q in questions}
    unknown = used_scopes - configured_scopes
    if unknown:
        sys.exit(f"ERROR: spreadsheet uses scope(s) not in CONFIG.SCOPES: {sorted(unknown)}\n"
                 f"       Add them (with a description) to the SCOPES list.")
    for grade in {q["grade"] for q in questions}:
        if grade not in grade_points:
            sys.exit(f"ERROR: question grade {grade!r} has no entry in grade_points")
    if not grading:
        sys.exit("ERROR: no grading thresholds were read from the Lookups sheet")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="Build the metric YAML from the scoring spreadsheet.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="path to the .xlsx template")
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="path to write the metric YAML")
    ap.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV, help="path to write the metric CSV")
    ap.add_argument("--check", action="store_true",
                    help="do not write; exit 1 if either output differs from what the spreadsheet would produce")
    args = ap.parse_args()

    if not args.workbook.exists():
        sys.exit(f"ERROR: workbook not found: {args.workbook}")

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    questions = read_questions(wb["Scoring"])
    grade_points = read_grade_points(wb["Lookups"])
    grading = read_grading(wb["Lookups"])
    validate(questions, grade_points, grading)

    generated_yaml = render_yaml(questions, grade_points, grading)
    generated_csv = render_csv(questions, grade_points)

    # Confirm the YAML output is itself valid before we trust it.
    yaml.safe_load(generated_yaml)

    # Compare/write as bytes so the CSV's CRLF line endings round-trip exactly
    # (read_text would silently translate them and break the comparison).
    targets = [(args.output, generated_yaml.encode("utf-8")),
               (args.output_csv, generated_csv.encode("utf-8"))]

    if args.check:
        drifted = [path for path, content in targets
                   if (path.read_bytes() if path.exists() else b"") != content]
        if not drifted:
            print(f"OK: {args.output.name} and {args.output_csv.name} are in sync "
                  f"with {args.workbook.name} ({len(questions)} questions)")
            sys.exit(0)
        for path in drifted:
            print(f"DRIFT: {path} differs from the spreadsheet.", file=sys.stderr)
        print(f"Run: python3 {SCRIPT_PATH}", file=sys.stderr)
        sys.exit(1)

    for path, content in targets:
        path.write_bytes(content)
    print(f"Wrote {args.output} and {args.output_csv} "
          f"({len(questions)} questions, "
          f"{sum(q['scope'] in NA_DEFAULT_SCOPES for q in questions)} ethics) "
          f"from {args.workbook.name}")


if __name__ == "__main__":
    main()
